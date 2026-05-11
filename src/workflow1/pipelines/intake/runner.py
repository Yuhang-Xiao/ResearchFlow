"""Minimal raw data intake and schema inventory runner."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import csv
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from workflow1.pipelines.runner import PipelineResult
from workflow1.utils.io import ensure_directory


SUPPORTED_RAW_EXTENSIONS = {".csv", ".xlsx"}


@dataclass(frozen=True)
class TableInventory:
    """Schema inventory for one file or workbook sheet."""

    dataset_name: str
    file_path: str
    file_type: str
    sheet_name: str
    row_count: int
    column_count: int
    columns: tuple[str, ...] = field(default_factory=tuple)
    notes: str = ""


def run(raw_dir: str | Path = "data/01_raw", reports_dir: str | Path = "reports") -> PipelineResult:
    """Scan raw files and write schema inventory outputs."""

    raw_path = Path(raw_dir)
    reports_path = Path(reports_dir)
    tables_path = ensure_directory(reports_path / "tables")
    ensure_directory(reports_path)

    raw_files = _list_raw_files(raw_path)
    if not raw_files:
        return PipelineResult(
            name="intake",
            status="no_raw_data",
            details={
                "message": f"No supported raw files found in {raw_path}. Add .csv or .xlsx files before running intake.",
                "raw_dir": str(raw_path),
                "supported_extensions": sorted(SUPPORTED_RAW_EXTENSIONS),
            },
        )

    inventories: list[TableInventory] = []
    warnings: list[str] = []
    for file_path in raw_files:
        try:
            inventories.extend(_inventory_file(file_path))
        except Exception as exc:  # pragma: no cover - defensive boundary for user files.
            warnings.append(f"{file_path}: {exc}")

    if not inventories:
        return PipelineResult(
            name="intake",
            status="error",
            details={"message": "Raw files were found, but no schema inventory could be created.", "warnings": warnings},
        )

    dataset_name = _dataset_name(raw_files)
    markdown_path = reports_path / f"schema_inventory_{dataset_name}.md"
    csv_path = tables_path / f"schema_inventory_{dataset_name}.csv"
    _write_inventory_markdown(markdown_path, inventories, warnings)
    _write_inventory_csv(csv_path, inventories)

    return PipelineResult(
        name="intake",
        status="ok",
        details={
            "dataset_name": dataset_name,
            "files": len(raw_files),
            "tables": len(inventories),
            "outputs": [str(markdown_path), str(csv_path)],
            "warnings": warnings,
        },
    )


def _list_raw_files(raw_path: Path) -> list[Path]:
    if not raw_path.exists():
        return []
    return sorted(
        path
        for path in raw_path.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_RAW_EXTENSIONS and not path.name.startswith("~$")
    )


def _inventory_file(file_path: Path) -> list[TableInventory]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _inventory_xlsx(file_path)
    if suffix == ".csv":
        return [_inventory_csv(file_path)]
    return []


def _inventory_xlsx(file_path: Path) -> list[TableInventory]:
    workbook = load_workbook(file_path, read_only=True, data_only=False)
    inventories: list[TableInventory] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        columns = tuple(_clean_column_name(value, index) for index, value in enumerate(header_row, start=1))
        row_count = max((sheet.max_row or 0) - 1, 0)
        column_count = sheet.max_column or len(columns)
        inventories.append(
            TableInventory(
                dataset_name=file_path.stem,
                file_path=str(file_path),
                file_type="xlsx",
                sheet_name=sheet_name,
                row_count=row_count,
                column_count=column_count,
                columns=columns,
            )
        )
    workbook.close()
    return inventories


def _inventory_csv(file_path: Path) -> TableInventory:
    encoding = _detect_csv_encoding(file_path)
    with file_path.open("r", encoding=encoding, errors="replace", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader, [])
        row_count = sum(1 for _ in reader)
    columns = tuple(_clean_column_name(value, index) for index, value in enumerate(header, start=1))
    return TableInventory(
        dataset_name=file_path.stem,
        file_path=str(file_path),
        file_type="csv",
        sheet_name="",
        row_count=row_count,
        column_count=len(columns),
        columns=columns,
        notes=f"encoding={encoding}",
    )


def _detect_csv_encoding(file_path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with file_path.open("r", encoding=encoding) as handle:
                handle.read(4096)
            return encoding
        except UnicodeDecodeError:
            continue
    return "utf-8"


def _clean_column_name(value: Any, index: int) -> str:
    if value is None or str(value).strip() == "":
        return f"Unnamed_{index}"
    return str(value).strip()


def _dataset_name(raw_files: list[Path]) -> str:
    if len(raw_files) == 1:
        return _safe_name(raw_files[0].stem)
    return "raw_inventory"


def _safe_name(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z_\-\u4e00-\u9fff]+", "_", value).strip("_") or "dataset"


def _write_inventory_csv(path: Path, inventories: list[TableInventory]) -> None:
    rows = [
        {
            "dataset_name": item.dataset_name,
            "file_path": item.file_path,
            "file_type": item.file_type,
            "sheet_name": item.sheet_name,
            "row_count": item.row_count,
            "column_count": item.column_count,
            "columns": " | ".join(item.columns),
            "notes": item.notes,
        }
        for item in inventories
    ]
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")


def _write_inventory_markdown(path: Path, inventories: list[TableInventory], warnings: list[str]) -> None:
    lines = [
        "# 原始数据 Schema Inventory",
        "",
        "本报告由最小 intake runner 生成，仅记录文件、sheet、行列规模和表头信息；未执行清洗、建模或可视化。",
        "",
        "| dataset_name | file_type | sheet_name | rows | columns |",
        "|---|---|---|---:|---:|",
    ]
    for item in inventories:
        lines.append(
            f"| `{item.dataset_name}` | `{item.file_type}` | `{item.sheet_name}` | {item.row_count} | {item.column_count} |"
        )
    lines.extend(["", "## 字段清单", ""])
    for item in inventories:
        title = item.dataset_name if not item.sheet_name else f"{item.dataset_name} / {item.sheet_name}"
        lines.extend([f"### {title}", ""])
        lines.extend(f"- `{column}`" for column in item.columns)
        if item.notes:
            lines.extend(["", f"备注：{item.notes}"])
        lines.append("")
    if warnings:
        lines.extend(["## 读取警告", ""])
        lines.extend(f"- {warning}" for warning in warnings)
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")

